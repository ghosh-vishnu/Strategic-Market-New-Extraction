from django.shortcuts import render

# Create your views here.
import os, uuid, threading, random, re, logging
from pathlib import Path
from typing import Set
from datetime import date

import pandas as pd
from django.conf import settings
from django.http import FileResponse, Http404, HttpResponseBadRequest
from django.utils.text import get_valid_filename
from openpyxl import load_workbook
from openpyxl.styles import Font
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response


from converter.utils import extractor
from converter.models import UploadedFile, JobRecord, ExcelMapping, ExtractExcelData
from django.utils import timezone
from django.db import DatabaseError

# Initialize logger
logger = logging.getLogger(__name__)

# simple in-memory job tracker
JOBS = {}
JOBS_LOCK = threading.Lock()

ALLOWED_DOCUMENT_EXTENSIONS = {'.doc', '.docx', '.rtf', '.odt'}


def _register_job_for_session(request, job_id: str) -> None:
    tracked = request.session.get("converter_jobs", [])
    if job_id not in tracked:
        tracked.append(job_id)
        request.session["converter_jobs"] = tracked
        request.session.modified = True


def _remove_job_from_session(request, job_id: str) -> None:
    tracked = request.session.get("converter_jobs", [])
    if job_id in tracked:
        tracked.remove(job_id)
        request.session["converter_jobs"] = tracked
        request.session.modified = True


def _job_is_authorized(request, job_id: str) -> bool:
    tracked = request.session.get("converter_jobs", [])
    return job_id in tracked


def _sanitize_filename(original_name: str, used_names: Set[str]) -> str:
    base_name = os.path.basename(original_name)
    base_name = get_valid_filename(base_name)

    if not base_name:
        base_name = f"document-{uuid.uuid4().hex}.docx"

    name_path = Path(base_name)
    stem = name_path.stem or "document"
    suffix = name_path.suffix or ".docx"

    candidate = f"{stem}{suffix}"
    counter = 1
    while candidate in used_names:
        candidate = f"{stem}_{counter}{suffix}"
        counter += 1

    used_names.add(candidate)
    return candidate

def _job_dir(job_id: str) -> Path:
    return Path(settings.MEDIA_ROOT) / job_id

def _cleanup_uploaded_files(folder: Path):
    """Clean up uploaded Word files after successful conversion, keeping only the output files."""
    try:
        for file_path in folder.iterdir():
            if file_path.is_file():
                # Keep only the output Excel and CSV files
                if not (file_path.name.endswith('.xlsx') or file_path.name.endswith('.csv')):
                    file_path.unlink()  # Delete the file
    except (OSError, IOError, PermissionError, FileNotFoundError) as e:
        # Log the error for debugging but don't raise - cleanup is not critical
        logger.warning(f"Could not clean up files in {folder}: {e}")
        return  # Exit quietly if cleanup fails

def _delete_job_folder(job_id: str):
    """Delete job folder and clean up from memory."""
    # Try to delete the folder
    try:
        folder = _job_dir(job_id)
        if folder.exists():
            import shutil
            shutil.rmtree(folder, ignore_errors=True)
    except (OSError, IOError, PermissionError, FileNotFoundError) as e:
        # Log folder deletion errors
        logger.warning(f"Could not delete job folder for {job_id}: {e}")
            
    # Try to remove from JOBS dictionary
    try:
        if job_id in JOBS:
            del JOBS[job_id]
    except KeyError as e:
        # Job already removed from memory
        logger.warning(f"Job {job_id} not found in JOBS dictionary")

def _cleanup_old_jobs():
    """Clean up old job folders that are no longer needed."""
    try:
        import time
        current_time = time.time()
        media_root = Path(settings.MEDIA_ROOT)
        
        for job_folder in media_root.iterdir():
            if job_folder.is_dir() and job_folder.name not in JOBS:
                try:
                    # Check if folder is older than 1 hour (3600 seconds)
                    folder_age = current_time - job_folder.stat().st_mtime
                    if folder_age > 3600:  # 1 hour
                        import shutil
                        shutil.rmtree(job_folder, ignore_errors=True)
                except (OSError, IOError, PermissionError, FileNotFoundError) as e:
                    # Skip individual folder errors
                    logger.warning(f"Could not clean up old folder {job_folder.name}: {e}")
    except (OSError, IOError, PermissionError, FileNotFoundError) as e:
        # Log media root access errors
        logger.warning(f"Error accessing media root during cleanup: {e}")

# ------------------- Upload API -------------------
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_files(request):
    # Clean up old job folders before starting new upload
    _cleanup_old_jobs()

    # Support batched uploads: if jobId provided, append files to same job
    incoming_job_id = request.GET.get("jobId")

    if incoming_job_id and incoming_job_id in JOBS:
        if not _job_is_authorized(request, incoming_job_id):
            return Response({"detail": "Job not found"}, status=status.HTTP_404_NOT_FOUND)
        job_id = incoming_job_id
        folder = _job_dir(job_id)
        folder.mkdir(parents=True, exist_ok=True)
        _register_job_for_session(request, job_id)
    else:
        job_id = str(uuid.uuid4())
        folder = _job_dir(job_id)
        folder.mkdir(parents=True, exist_ok=True)
        
        # Create job record in database
        JobRecord.objects.create(
            job_id=job_id,
            folder_name="Word_Files",  # Will be updated below
            progress=0,
            status="pending",
            is_active=True
        )
        _register_job_for_session(request, job_id)

    files = request.FILES.getlist('files')
    if not files:
        return HttpResponseBadRequest("No files uploaded")

    invalid_files = [
        f.name for f in files
        if Path(f.name).suffix.lower() not in ALLOWED_DOCUMENT_EXTENSIONS
    ]
    if invalid_files:
        return Response(
            {
                "detail": "Unsupported file type(s) detected.",
                "files": invalid_files,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    existing_names: Set[str] = set()
    if folder.exists():
        existing_names.update(os.listdir(folder))

    # Extract folder name only when creating a NEW job
    if incoming_job_id is None or incoming_job_id not in JOBS:
        folder_name = "Word_Files"  # Default name
        if files and hasattr(files[0], 'name'):
            # Try to get folder name from webkitRelativePath if available
            webkit_path = getattr(files[0], 'webkitRelativePath', None)
            
            if webkit_path and '/' in webkit_path:
                folder_name = webkit_path.split('/')[0]
            elif webkit_path and '\\' in webkit_path:
                folder_name = webkit_path.split('\\')[0]
            else:
                # If no webkitRelativePath, use a sanitized version of the first file's directory
                folder_name = "Word_Files"

        # Sanitize folder name for filename
        import re
        original_folder_name = folder_name
        folder_name = re.sub(r'[^\w\s-]', '', folder_name).strip()
        folder_name = re.sub(r'[-\s]+', '_', folder_name)
        if not folder_name:
            folder_name = "Word_Files"
        
    else:
        folder_name = JOBS[job_id].get("folder_name", "Word_Files")

    # Save file information to database
    uploaded_files = []
    for f in files:
        filename = f.name
        safe_name = _sanitize_filename(filename, existing_names)
        path = folder / safe_name
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, 'wb+') as dest:
            for chunk in f.chunks():
                dest.write(chunk)
        
        # Save file info to database
        uploaded_file = UploadedFile.objects.create(
            job_id=job_id,
            folder_name=folder_name,
            file_path=str(path),
            file_name=safe_name,
            file_size=f.size,
            upload_date=timezone.now(),
            conversion_complete=False,
            download_complete=False
        )
        uploaded_files.append(uploaded_file)

    if incoming_job_id is None or incoming_job_id not in JOBS:
        JOBS[job_id] = {"progress": 0, "done": False, "result": None, "error": None, "folder_name": folder_name, "cancelled": False}
        # Update job record in database with folder name and file count
        try:
            job_record = JobRecord.objects.get(job_id=job_id)
            job_record.folder_name = folder_name
            job_record.save()
        except JobRecord.DoesNotExist:
            logger.warning(f"Job record not found for {job_id}")
    # For batched appends, keep existing JOBS entry
    return Response({"jobId": job_id})

# ------------------- Worker function -------------------

def _convert_worker(job_id: str):
    try:
        JOBS[job_id]["progress"] = 5
        JOBS[job_id]["status_message"] = "Initializing conversion..."
        folder = _job_dir(job_id)

        files_to_process = [f for f in os.listdir(folder) if f.endswith(".docx") and not f.startswith("~$")]
        total_files = len(files_to_process)
        
        if total_files == 0:
            JOBS[job_id]["progress"] = 100
            JOBS[job_id]["done"] = True
            JOBS[job_id]["status_message"] = "No Word files found to process"
            return

        all_data = []
        JOBS[job_id]["status_message"] = f"Processing {total_files} files..."
        
        for i, file in enumerate(files_to_process):
            if JOBS.get(job_id, {}).get("cancelled"):
                JOBS[job_id]["error"] = "cancelled"
                JOBS[job_id]["done"] = True
                JOBS[job_id]["status_message"] = "Conversion cancelled"
                _cleanup_uploaded_files(folder)
                return

            path = folder / file
            logger.info(f"Processing {file}... ({i+1}/{total_files})")

            # extract fields
            title = extractor.extract_title(str(path))
            description = extractor.extract_description(str(path))
            toc = extractor.extract_toc(str(path))
            methodology = extractor.extract_methodology_from_faqschema(str(path))
            seo_title = extractor.extract_seo_title(str(path))
            breadcrumb_text = extractor.extract_breadcrumb_text(str(path))
            skucode = extractor.extract_sku_code(str(path))
            urlrp = extractor.extract_sku_code(str(path))
            breadcrumb_schema = extractor.extract_breadcrumb_schema(str(path))
            meta = extractor.extract_meta_description(str(path))
            schema2 = extractor.extract_faq_schema(str(path))
            report = extractor.extract_report_coverage_table_with_style(str(path))

            # ✅ merge description + report
            merged_text = (description or "") + "\n\n" + (report or "")

            # ✅ split into parts
            chunks = extractor.split_into_excel_cells(merged_text)

            row_data = {
                "File": file,
                "Title": title,
            }

            # add merged description parts
            for j, chunk in enumerate(chunks, start=1):
                row_data[f"Description_Part{j}"] = chunk

            # add other fields (without Report, because merged already)
            row_data.update({
                "TOC": toc,
                "Segmentation": "<p>.</p>",
                "Methodology": methodology,
                "Publish_Date": date.today().strftime('%b-%Y').upper(),
                "Image": "",  # Blank image column
                "Currency": "USD",
                "Single Price": 4485,
                "Corporate Price": 6449,
                "skucode": skucode,
                "Total Page": random.randint(150, 200),
                "RID": "",  # RID column after Total Page
                "Date": date.today().strftime("%d-%m-%Y"),
                "Status": "IN",  # Default status
                "Report_Docs": "",  # Report docs column
                "urlNp": urlrp,
                "Meta Description": meta,
                "Meta_Key": ".",  # Meta key with dot
                "Base Year": "2024",
                "history": "2019-2023",
                "Enterprise Price": 8339,
                "SEOTITLE": seo_title,
                "BreadCrumb Text": breadcrumb_text,
                "Schema 1": breadcrumb_schema,
                "Schema 2": schema2,
                "Sub-Category": ""  # Sub-Category column
                # ⚠ Report removed
            })
            
            all_data.append(row_data)
            
            # Update progress after each file (80% for file processing, 20% for final steps)
            file_progress = 5 + int((i + 1) / total_files * 80)
            # Ensure progress never goes backwards
            if file_progress > JOBS[job_id]["progress"]:
                JOBS[job_id]["progress"] = file_progress

        if JOBS.get(job_id, {}).get("cancelled"):
            JOBS[job_id]["error"] = "cancelled"
            JOBS[job_id]["done"] = True
            JOBS[job_id]["status_message"] = "Conversion cancelled"
            _cleanup_uploaded_files(folder)
            return

        JOBS[job_id]["status_message"] = "Creating Excel file..."
        df = pd.DataFrame(all_data)

        # enforce column order
        desc_parts = sorted([c for c in df.columns if c.startswith("Description_Part")],
                            key=lambda x: int(x.replace("Description_Part", "")))

        # Separate Description_Part1 and other Description_Parts
        desc_part1 = [c for c in desc_parts if c == "Description_Part1"]
        other_desc_parts = [c for c in desc_parts if c != "Description_Part1"]

        columns_order = ["File", "Title"] + desc_part1 + [
            "TOC", "Segmentation", "Methodology", "Publish_Date", "Image", "Currency",
            "Single Price",  "Corporate Price", "skucode", "Total Page", "RID", "Date", "Status", "Report_Docs",
            "urlNp", "Meta Description", "Meta_Key", "Base Year", "history",
            "Enterprise Price", "SEOTITLE", "BreadCrumb Text", "Schema 1", "Schema 2", "Sub-Category"
        ] + other_desc_parts  # Add other Description_Parts at the end

        df = df[[col for col in columns_order if col in df.columns]]

        folder_name = JOBS[job_id].get("folder_name", "Word_Files")
        xlsx_path = folder / f"{folder_name}.xlsx"
        csv_path = folder / f"{folder_name}.csv"

        df.to_excel(xlsx_path, index=False)
        JOBS[job_id]["status_message"] = "Applying formatting..."
        
        # Apply bold formatting to Publish_Date column
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # Find Publish_Date column index
        publish_date_col = None
        for col_idx, header in enumerate(ws[1], 1):
            if header.value == "Publish_Date":
                publish_date_col = col_idx
                break
        
        # Apply bold formatting to Publish_Date column
        if publish_date_col:
            bold_font = Font(bold=True)
            for row in range(2, ws.max_row + 1):  # Skip header row
                cell = ws.cell(row=row, column=publish_date_col)
                cell.font = bold_font
        
        wb.save(xlsx_path)
        JOBS[job_id]["status_message"] = "Creating CSV file..."
        
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")

        # Ensure paths are absolute
        xlsx_abs_path = str(xlsx_path.absolute())
        csv_abs_path = str(csv_path.absolute())
        
        JOBS[job_id]["result"] = {"xlsx": xlsx_abs_path, "csv": csv_abs_path}
        JOBS[job_id]["progress"] = 100
        JOBS[job_id]["done"] = True
        JOBS[job_id]["status_message"] = "Conversion complete!"
        

        # Mark files as conversion complete in database
        UploadedFile.objects.filter(job_id=job_id).update(conversion_complete=True)
        
        # Update job record in database
        try:
            job_record = JobRecord.objects.get(job_id=job_id)
            job_record.progress = 100
            job_record.status = "completed"
            job_record.save()
        except JobRecord.DoesNotExist:
            logger.warning(f"Job record not found for {job_id}")

        _cleanup_uploaded_files(folder)
        
        # Clean up database automatically after conversion completes
        import threading
        def cleanup_job_db():
            """Automatically clean up database after conversion completes"""
            try:
                # Delete uploaded files
                UploadedFile.objects.filter(job_id=job_id).delete()
                # Delete Excel mapping data
                ExcelMapping.objects.filter(job_id=job_id).delete()
                # Delete Extract Excel data
                ExtractExcelData.objects.filter(job_id=job_id).delete()
                # Delete job record
                JobRecord.objects.filter(job_id=job_id).delete()
                logger.info(f"Automatically cleaned up database for job {job_id}")
            except Exception as e:
                logger.warning(f"Could not clean up database for job {job_id}: {e}")
        
        def cleanup_job_memory():
            import time
            time.sleep(300)  # Wait 5 minutes before cleaning up from memory
            if job_id in JOBS:
                with JOBS_LOCK:
                    if job_id in JOBS:
                        del JOBS[job_id]
                        logger.info(f"Cleaned up completed job {job_id} from memory")
        
        # Start background cleanup tasks
        db_cleanup_thread = threading.Thread(target=cleanup_job_db, daemon=True)
        db_cleanup_thread.start()
        
        memory_cleanup_thread = threading.Thread(target=cleanup_job_memory, daemon=True)
        memory_cleanup_thread.start()

    except Exception as e:
        logger.error(f"Error in conversion worker for job {job_id}: {str(e)}", exc_info=True)
        JOBS[job_id]["error"] = str(e)
        JOBS[job_id]["done"] = True
        _cleanup_uploaded_files(folder)
        
        # Clean up the job from memory after some time
        import threading
        def cleanup_job():
            import time
            time.sleep(300)  # Wait 5 minutes before cleaning up
            if job_id in JOBS:
                del JOBS[job_id]
                print(f"Cleaned up job {job_id} from memory")
        
        cleanup_thread = threading.Thread(target=cleanup_job, daemon=True)
        cleanup_thread.start()



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reset_job(request):
    """Cancel a running job and delete its files. If jobId missing, clear all jobs."""
    job_id = request.GET.get("jobId")
    if job_id:
        if not _job_is_authorized(request, job_id):
            return Response({"reset": False, "detail": "Job not found"}, status=status.HTTP_404_NOT_FOUND)
        # Clean up database records first (in case job was already deleted from memory)
        try:
            # Delete uploaded files
            UploadedFile.objects.filter(job_id=job_id).delete()
            # Delete Excel mapping data
            ExcelMapping.objects.filter(job_id=job_id).delete()
            # Delete Extract Excel data
            ExtractExcelData.objects.filter(job_id=job_id).delete()
            # Delete job record entirely
            JobRecord.objects.filter(job_id=job_id).delete()
        except (DatabaseError, Exception) as e:
            # Database cleanup errors are not critical - records may already be deleted
            logger.warning(f"Could not clean up database for job {job_id}: {e}")
            # Continue with memory cleanup even if DB cleanup fails
        
        # Clean up memory only if job exists
        try:
            if job_id in JOBS:
                JOBS[job_id]["cancelled"] = True
                _delete_job_folder(job_id)
                with JOBS_LOCK:
                    if job_id in JOBS:
                        del JOBS[job_id]
        except KeyError:
            # Job was already removed from memory
            logger.info(f"Job {job_id} was already removed from memory")

        _remove_job_from_session(request, job_id)
        
        return Response({"reset": True, "jobId": job_id})
    # No jobId: clear all
    for jid in list(JOBS.keys()):
        try:
            if jid in JOBS:
                JOBS[jid]["cancelled"] = True
                _delete_job_folder(jid)
                with JOBS_LOCK:
                    if jid in JOBS:
                        del JOBS[jid]
        except KeyError:
            # Job was already removed by another thread/process
            logger.info(f"Job {jid} was already removed during clear all")
    request.session["converter_jobs"] = []
    request.session.modified = True
    return Response({"reset": True, "all": True})

# ------------------- Start Conversion -------------------
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def start_convert(request):
    job_id = request.GET.get("jobId")
    if not job_id or job_id not in JOBS:
        return Response({"detail": "Invalid jobId"}, status=status.HTTP_400_BAD_REQUEST)
    if not _job_is_authorized(request, job_id):
        return Response({"detail": "Job not found"}, status=status.HTTP_404_NOT_FOUND)

    # Update job status in database to "converting"
    try:
        job_record = JobRecord.objects.get(job_id=job_id)
        job_record.status = "converting"
        job_record.progress = 5
        job_record.save()
    except JobRecord.DoesNotExist:
        print(f"WARNING: Job record not found for {job_id}")

    t = threading.Thread(target=_convert_worker, args=(job_id,), daemon=True)
    t.start()
    return Response({"started": True})

# ------------------- Progress -------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def progress(request):
    job_id = request.GET.get("jobId")
    if not job_id:
        return Response({"error": "Job not found", "progress": 0, "done": False}, status=404)

    job_data = JOBS.get(job_id)
    if not job_data or not _job_is_authorized(request, job_id):
        job_record = JobRecord.objects.filter(job_id=job_id).first()
        if job_record:
            return Response(
                {
                    "progress": job_record.progress or 0,
                    "done": job_record.status == "completed",
                    "error": "Job state was reset. Please restart the conversion.",
                },
                status=status.HTTP_409_CONFLICT,
            )
        return Response({"error": "Job not found", "progress": 0, "done": False}, status=404)
    
    job_data = job_data

    # Check if job has an error
    if job_data.get("error"):
        return Response({
            "error": job_data["error"],
            "progress": job_data.get("progress", 0),
            "done": job_data.get("done", False)
        })
    
    # Include status message in response
    response_data = {
        "progress": job_data.get("progress", 0),
        "done": job_data.get("done", False),
        "error": job_data.get("error"),
        "status_message": job_data.get("status_message", "Processing...")
    }
    
    return Response(response_data)

# ------------------- Extract Excel Upload -------------------
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_extract_excel(request):
    """Upload Extract Excel sheet"""
    try:
        logger.info(f"DEBUG: Extract Excel upload request received")
        logger.info(f"DEBUG: POST data: {request.POST}")
        logger.info(f"DEBUG: FILES data: {request.FILES}")
        
        job_id = request.POST.get("jobId")
        logger.info(f"DEBUG: Job ID from request: {job_id}")
        
        if not job_id or job_id not in JOBS:
            logger.info(f"DEBUG: Invalid jobId: {job_id}, Available jobs: {list(JOBS.keys())}")
            return Response({"success": False, "message": "Invalid jobId"}, status=400)
        if not _job_is_authorized(request, job_id):
            return Response({"success": False, "message": "Job not found"}, status=404)
        
        excel_file = request.FILES.get("excelFile")
        logger.info(f"DEBUG: Excel file: {excel_file}")
        
        if not excel_file:
            return Response({"success": False, "message": "No Excel file provided"}, status=400)
        
        # Validate file extension
        logger.info(f"DEBUG: File name: '{excel_file.name}'")
        logger.info(f"DEBUG: File name lower: '{excel_file.name.lower()}'")
        logger.info(f"DEBUG: File name length: {len(excel_file.name)}")
        
        # Check file extension more carefully
        file_name_lower = excel_file.name.lower().strip()
        if not (file_name_lower.endswith('.xlsx') or file_name_lower.endswith('.xls')):
            logger.info(f"DEBUG: File extension validation failed for: '{file_name_lower}'")
            return Response({"success": False, "message": "Please upload an Excel file (.xlsx or .xls)"}, status=400)
        logger.info(f"DEBUG: File extension validation passed")
        
        # Read Excel file
        import pandas as pd
        try:
            logger.info(f"DEBUG: Reading Excel file...")
            df = pd.read_excel(excel_file)
            logger.info(f"DEBUG: Excel file read successfully. Shape: {df.shape}")
            logger.info(f"DEBUG: Columns: {df.columns.tolist()}")
        except Exception as e:
            logger.info(f"DEBUG: Error reading Excel file: {str(e)}")
            return Response({"success": False, "message": f"Error reading Excel file: {str(e)}"}, status=400)
        
        # Store extract data in database
        import json
        import numpy as np
        from datetime import date, datetime
        
        # Helper function to convert numpy types to Python native types
        def convert_to_native(obj):
            if isinstance(obj, (np.integer, np.int64, np.int32)):
                return int(obj)
            elif isinstance(obj, (np.floating, np.float64, np.float32)):
                return float(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif pd.isna(obj):
                return None
            elif isinstance(obj, (date, datetime)):
                return obj.isoformat()
            else:
                return str(obj) if obj is not None else None
        
        records = df.to_dict('records')
        for record in records:
            # Convert all values to Python native types
            cleaned_record = {k: convert_to_native(v) for k, v in record.items()}
            ExtractExcelData.objects.create(
                job_id=job_id,
                row_data=cleaned_record
            )
        
        # Also store in job for in-memory access
        JOBS[job_id]['extract_excel_data'] = records
        JOBS[job_id]['extract_excel_uploaded'] = True
        
        logger.info(f"DEBUG: Extract Excel uploaded successfully for job {job_id} with {len(df)} entries")
        logger.info(f"DEBUG: Extract Excel data saved to database for {len(records)} entries")
        
        return Response({
            "success": True,
            "message": "Extract Excel sheet uploaded successfully",
            "entries": len(df)
        })
        
    except Exception as e:
        print(f"ERROR: Extract Excel upload failed: {str(e)}")
        return Response({"success": False, "message": f"Upload failed: {str(e)}"}, status=500)

# ------------------- Direct Excel Upload (Skip Word Conversion) -------------------
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_direct_excel(request):
    """Upload converted Excel file directly (skip Word conversion step)"""
    try:
        logger.info(f"DEBUG: Direct Excel upload request received")
        logger.info(f"DEBUG: POST data: {request.POST}")
        logger.info(f"DEBUG: FILES data: {request.FILES}")
        
        # Create a new job for direct Excel upload
        job_id = str(uuid.uuid4())
        job_dir = _job_dir(job_id)
        job_dir.mkdir(parents=True, exist_ok=True)
        _register_job_for_session(request, job_id)
        
        excel_file = request.FILES.get("excelFile")
        logger.info(f"DEBUG: Excel file: {excel_file}")
        
        if not excel_file:
            return Response({"success": False, "message": "No Excel file provided"}, status=400)
        
        # Save the Excel file
        excel_path = job_dir / excel_file.name
        with open(excel_path, 'wb') as f:
            for chunk in excel_file.chunks():
                f.write(chunk)
        
        logger.info(f"DEBUG: Excel file saved to: {excel_path}")
        
        # Read the Excel file to get basic info
        try:
            df = pd.read_excel(excel_path)
            logger.info(f"DEBUG: Excel file read successfully. Shape: {df.shape}")
            logger.info(f"DEBUG: Columns: {list(df.columns)}")
        except Exception as e:
            print(f"ERROR: Failed to read Excel file: {str(e)}")
            return Response({"success": False, "message": f"Invalid Excel file: {str(e)}"}, status=400)
        
        # Store extract data in database
        import json
        import numpy as np
        from datetime import date, datetime
        
        # Helper function to convert numpy types to Python native types
        def convert_to_native(obj):
            if isinstance(obj, (np.integer, np.int64, np.int32)):
                return int(obj)
            elif isinstance(obj, (np.floating, np.float64, np.float32)):
                return float(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif pd.isna(obj):
                return None
            elif isinstance(obj, (date, datetime)):
                return obj.isoformat()
            else:
                return str(obj) if obj is not None else None
        
        records = df.to_dict('records')
        for record in records:
            # Convert all values to Python native types
            cleaned_record = {k: convert_to_native(v) for k, v in record.items()}
            ExtractExcelData.objects.create(
                job_id=job_id,
                row_data=cleaned_record
            )
        logger.info(f"DEBUG: Extract Excel data saved to database for {len(records)} entries")
        
        # Create job record in database
        JobRecord.objects.create(
            job_id=job_id,
            folder_name="Direct_Excel_Upload",
            progress=100,
            status="completed",
            is_active=True
        )
        
        # Initialize job data
        JOBS[job_id] = {
            "status": "excel_uploaded",
            "excel_uploaded": True,
            "excel_path": str(excel_path),
            "excel_filename": excel_file.name,
            "extract_excel_data": records,  # Store in memory for quick access
            "extract_excel_uploaded": True,
            "done": True,  # Mark as done since we're skipping conversion
            "result": {
                "xlsx": str(excel_path)
            },
            "created_at": date.today().isoformat()
        }
        
        logger.info(f"DEBUG: Job {job_id} initialized for direct Excel upload")
        
        return Response({
            "success": True, 
            "message": "Excel file uploaded successfully",
            "jobId": job_id,
            "filename": excel_file.name,
            "entries": len(df)
        })
        
    except Exception as e:
        print(f"ERROR: Direct Excel upload failed: {str(e)}")
        return Response({"success": False, "message": f"Upload failed: {str(e)}"}, status=500)

# ------------------- Excel Upload -------------------
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_excel_sheet(request):
    """Upload Excel sheet with Title, Category, Subcategory columns"""
    try:
        logger.info(f"DEBUG: Excel upload request received")
        logger.info(f"DEBUG: POST data: {request.POST}")
        logger.info(f"DEBUG: FILES data: {request.FILES}")
        
        job_id = request.POST.get("jobId")
        logger.info(f"DEBUG: Job ID from request: {job_id}")
        
        if not job_id or job_id not in JOBS:
            logger.info(f"DEBUG: Invalid jobId: {job_id}, Available jobs: {list(JOBS.keys())}")
            return Response({"success": False, "message": "Invalid jobId"}, status=400)
        if not _job_is_authorized(request, job_id):
            return Response({"success": False, "message": "Job not found"}, status=404)
        
        excel_file = request.FILES.get("excelFile")
        logger.info(f"DEBUG: Excel file: {excel_file}")
        
        if not excel_file:
            return Response({"success": False, "message": "No Excel file provided"}, status=400)
        
        # Validate file extension
        logger.info(f"DEBUG: File name: '{excel_file.name}'")
        logger.info(f"DEBUG: File name lower: '{excel_file.name.lower()}'")
        logger.info(f"DEBUG: File name length: {len(excel_file.name)}")
        
        # Check file extension more carefully
        file_name_lower = excel_file.name.lower().strip()
        if not (file_name_lower.endswith('.xlsx') or file_name_lower.endswith('.xls')):
            logger.info(f"DEBUG: File extension validation failed for: '{file_name_lower}'")
            return Response({"success": False, "message": "Please upload an Excel file (.xlsx or .xls)"}, status=400)
        logger.info(f"DEBUG: File extension validation passed")
        
        # Read Excel file
        import pandas as pd
        try:
            logger.info(f"DEBUG: Reading Excel file...")
            df = pd.read_excel(excel_file)
            logger.info(f"DEBUG: Excel file read successfully. Shape: {df.shape}")
            logger.info(f"DEBUG: Columns: {df.columns.tolist()}")
        except Exception as e:
            logger.info(f"DEBUG: Error reading Excel file: {str(e)}")
            return Response({"success": False, "message": f"Error reading Excel file: {str(e)}"}, status=400)
        
        # Validate required columns - handle both 'Subcategory' and 'Sub-Category'
        required_columns = ['Title', 'Category']
        subcategory_columns = ['Subcategory', 'Sub-Category']
        subcategory_url_columns = ['Sub-Category Url', 'Sub-Category URL', 'Subcategory Url', 'Subcategory URL']
        
        logger.info(f"DEBUG: Required columns: {required_columns}")
        logger.info(f"DEBUG: Available columns: {df.columns.tolist()}")
        
        # Check for required columns
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        # Check for subcategory column (either 'Subcategory' or 'Sub-Category')
        subcategory_col = None
        for col in subcategory_columns:
            if col in df.columns:
                subcategory_col = col
                break
        
        if not subcategory_col:
            missing_columns.append('Subcategory or Sub-Category')
        
        # Check for subcategory URL column
        subcategory_url_col = None
        for col in subcategory_url_columns:
            if col in df.columns:
                subcategory_url_col = col
                break
        
        if not subcategory_url_col:
            missing_columns.append('Sub-Category Url')
        
        if missing_columns:
            logger.info(f"DEBUG: Missing columns: {missing_columns}")
            return Response({
                "success": False, 
                "message": f"Missing required columns: {', '.join(missing_columns)}. Required columns: Title, Category, Subcategory (or Sub-Category), Sub-Category Url"
            }, status=400)
        
        logger.info(f"DEBUG: All required columns found. Subcategory column: {subcategory_col}, Sub-Category Url column: {subcategory_url_col}")
        
        # Convert to list of dictionaries
        mapping_data = []
        for _, row in df.iterrows():
            mapping_data.append({
                'title': str(row['Title']).strip(),
                'category': str(row['Category']).strip(),
                'subcategory': str(row[subcategory_col]).strip(),
                'subcategory_url': str(row[subcategory_url_col]).strip() if pd.notna(row[subcategory_url_col]) else ''
            })
        
        # Store mapping data in database
        for entry in mapping_data:
            ExcelMapping.objects.create(
                job_id=job_id,
                title=entry['title'],
                category=entry['category'],
                subcategory=entry['subcategory'],
                subcategory_url=entry['subcategory_url']
            )
        
        # Also store in job for in-memory access
        JOBS[job_id]['excel_mapping'] = mapping_data
        JOBS[job_id]['excel_uploaded'] = True
        
        logger.info(f"DEBUG: Excel uploaded successfully for job {job_id} with {len(mapping_data)} entries")
        logger.info(f"DEBUG: Excel mapping saved to database for {len(mapping_data)} entries")
        
        return Response({
            "success": True,
            "message": "Excel sheet uploaded successfully",
            "entries": len(mapping_data)
        })
        
    except Exception as e:
        print(f"ERROR: Excel upload failed: {str(e)}")
        return Response({"success": False, "message": f"Upload failed: {str(e)}"}, status=500)

# ------------------- Apply Excel Mapping -------------------
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def apply_excel_mapping(request):
    """Apply Excel mapping to the converted Word files"""
    try:
        job_id = request.POST.get("jobId")
        if not job_id or job_id not in JOBS:
            return Response({"success": False, "message": "Invalid jobId"}, status=400)
        if not _job_is_authorized(request, job_id):
            return Response({"success": False, "message": "Job not found"}, status=404)
        
        # Check if either mapping Excel is uploaded OR direct Excel is uploaded
        if not JOBS[job_id].get('excel_uploaded', False) and not JOBS[job_id].get('excel_path'):
            return Response({"success": False, "message": "No mapping Excel sheet uploaded"}, status=400)
        
        
        # Check if mapping data exists (either from mapping Excel or needs to be created for direct Excel)
        if 'excel_mapping' not in JOBS[job_id]:
            # For direct Excel upload, we need to create a basic mapping
            if JOBS[job_id].get('excel_path'):
                # Create a basic mapping from the direct Excel file
                try:
                    df = pd.read_excel(JOBS[job_id]['excel_path'])
                    mapping_data = []
                    for _, row in df.iterrows():
                        # Use filename as title if Title column doesn't exist
                        title = str(row.get('Title', '')).strip() if 'Title' in df.columns else ''
                        mapping_data.append({
                            'title': title,
                            'category': str(row.get('Category', '')).strip() if 'Category' in df.columns else '',
                            'subcategory': str(row.get('Subcategory', row.get('Sub-Category', ''))).strip() if 'Subcategory' in df.columns or 'Sub-Category' in df.columns else '',
                            'subcategory_url': str(row.get('Sub-Category Url', '')).strip() if 'Sub-Category Url' in df.columns else ''
                        })
                    JOBS[job_id]['excel_mapping'] = mapping_data
                    logger.info(f"DEBUG: Created basic mapping for direct Excel upload with {len(mapping_data)} entries")
                except Exception as e:
                    return Response({"success": False, "message": f"Error creating mapping from direct Excel: {str(e)}"}, status=400)
            else:
                return Response({"success": False, "message": "Excel mapping data not found"}, status=400)
        
        # Check if conversion is complete (for regular flow) or if direct Excel is uploaded
        if not JOBS[job_id].get('done', False) and not JOBS[job_id].get('excel_path'):
            return Response({"success": False, "message": "Word files conversion not complete yet"}, status=400)
        
        # Get the result file path - check for direct Excel upload first, then regular result
        result_path = JOBS[job_id].get("excel_path") or JOBS[job_id].get("result", {}).get("xlsx")
        if not result_path or not os.path.exists(result_path):
            return Response({"success": False, "message": "Result file not found"}, status=400)
        
        # Read the converted Excel file
        import pandas as pd
        try:
            df = pd.read_excel(result_path)
        except Exception as e:
            return Response({"success": False, "message": f"Error reading result file: {str(e)}"}, status=400)
        
        mapping_data = JOBS[job_id]['excel_mapping']
        
        # Create mapping dictionary
        title_mapping = {}
        for entry in mapping_data:
            title_mapping[entry['title']] = {
                'title': entry['title'],  # Include title in mapping
                'category': entry['category'],
                'subcategory': entry['subcategory'],
                'subcategory_url': entry['subcategory_url']
            }
        
        logger.info(f"DEBUG: Applying mapping with {len(title_mapping)} entries")
        
        # Apply mapping to DataFrame
        if 'File' in df.columns:
            logger.info(f"DEBUG: Before mapping - DataFrame columns: {df.columns.tolist()}")
            logger.info(f"DEBUG: Before mapping - First few File values: {df['File'].head().tolist()}")
            
            # Keep Title column and add Category column after it
            if 'Title' not in df.columns:
                logger.info(f"DEBUG: Title column not found, adding it")
                df['Title'] = ''
            else:
                logger.info(f"DEBUG: Title column already exists with values: {df['Title'].head().tolist()}")
                # Preserve original Title values - don't overwrite them
            
            # Simple logic: Sub-Category already exists, just add Category column
            logger.info(f"DEBUG: Sub-Category column already exists: True")
            logger.info(f"DEBUG: Using existing Sub-Category column")
            
            # Add Category and Sub-Category Url columns (keep Title)
            df['Category'] = ''
            df['Sub-Category Url'] = ''
            logger.info(f"DEBUG: Added Category and Sub-Category Url columns, keeping Title column")
            
            # Reorder columns to have Category first, then Title
            columns = df.columns.tolist()
            logger.info(f"DEBUG: Before reordering - columns: {columns}")
            
            # Create new column order: File, Category, Title, then rest
            new_columns = []
            
            # Add File column first if it exists
            if 'File' in columns:
                new_columns.append('File')
                columns.remove('File')
            
            # Add Category column second
            if 'Category' in columns:
                new_columns.append('Category')
                columns.remove('Category')
            
            # Add Title column third
            if 'Title' in columns:
                new_columns.append('Title')
                columns.remove('Title')
            
            # Add Sub-Category column fourth
            if 'Sub-Category' in columns:
                new_columns.append('Sub-Category')
                columns.remove('Sub-Category')
            
            # Add Sub-Category Url column fifth
            if 'Sub-Category Url' in columns:
                new_columns.append('Sub-Category Url')
                columns.remove('Sub-Category Url')
            
            # Add remaining columns
            new_columns.extend(columns)
            
            df = df[new_columns]
            logger.info(f"DEBUG: After reordering - columns: {df.columns.tolist()}")
            
            # Apply mapping
            mapped_count = 0
            for idx, row in df.iterrows():
                # Extract filename from path and remove .docx extension
                file_path = str(row['File'])
                filename = file_path.split('/')[-1].replace('.docx', '').strip()
                logger.info(f"DEBUG: Processing file: '{filename}' (original: '{file_path}')")
                
                # Try exact match first
                if filename in title_mapping:
                    # Only update Title if it's empty or if mapping provides a better title
                    current_title = df.at[idx, 'Title'] if 'Title' in df.columns else ''
                    title_value = title_mapping[filename]['title']
                    category_value = title_mapping[filename]['category']
                    subcategory_value = title_mapping[filename]['subcategory']
                    subcategory_url_value = title_mapping[filename]['subcategory_url']
                    
                    # Update Title only if current title is empty or mapping title is more meaningful
                    if not current_title or current_title.strip() == '' or len(title_value.strip()) > len(current_title.strip()):
                        df.at[idx, 'Title'] = title_value
                        logger.info(f"DEBUG: Updated Title from '{current_title}' to '{title_value}'")
                    else:
                        logger.info(f"DEBUG: Keeping original Title: '{current_title}' (mapping title: '{title_value}')")
                    
                    df.at[idx, 'Category'] = category_value
                    df.at[idx, 'Sub-Category'] = subcategory_value
                    df.at[idx, 'Sub-Category Url'] = subcategory_url_value
                    mapped_count += 1
                    logger.info(f"DEBUG: Mapped {filename} -> Title: '{df.at[idx, 'Title']}', Category: '{category_value}', Sub-Category: '{subcategory_value}', Sub-Category Url: '{subcategory_url_value}'")
                else:
                    # Try case-insensitive match
                    found_match = False
                    for excel_title in title_mapping.keys():
                        if excel_title.lower().strip() == filename.lower().strip():
                            # Only update Title if it's empty or if mapping provides a better title
                            current_title = df.at[idx, 'Title'] if 'Title' in df.columns else ''
                            title_value = title_mapping[excel_title]['title']
                            category_value = title_mapping[excel_title]['category']
                            subcategory_value = title_mapping[excel_title]['subcategory']
                            subcategory_url_value = title_mapping[excel_title]['subcategory_url']
                            
                            # Update Title only if current title is empty or mapping title is more meaningful
                            if not current_title or current_title.strip() == '' or len(title_value.strip()) > len(current_title.strip()):
                                df.at[idx, 'Title'] = title_value
                                logger.info(f"DEBUG: Updated Title from '{current_title}' to '{title_value}'")
                            else:
                                logger.info(f"DEBUG: Keeping original Title: '{current_title}' (mapping title: '{title_value}')")
                            
                            df.at[idx, 'Category'] = category_value
                            df.at[idx, 'Sub-Category'] = subcategory_value
                            df.at[idx, 'Sub-Category Url'] = subcategory_url_value
                            mapped_count += 1
                            found_match = True
                            logger.info(f"DEBUG: Case-insensitive mapped {filename} -> {excel_title} -> Title: '{df.at[idx, 'Title']}', Category: '{category_value}', Sub-Category: '{subcategory_value}', Sub-Category Url: '{subcategory_url_value}'")
                            break
                    
                    if not found_match:
                        logger.info(f"DEBUG: No mapping found for: '{filename}'")
                        logger.info(f"DEBUG: Available mapping keys: {list(title_mapping.keys())[:5]}")
                        logger.info(f"DEBUG: Trying to find partial matches...")
                        
                        # Try partial match (contains)
                        for excel_title in title_mapping.keys():
                            if filename.lower() in excel_title.lower() or excel_title.lower() in filename.lower():
                                # Only update Title if it's empty or if mapping provides a better title
                                current_title = df.at[idx, 'Title'] if 'Title' in df.columns else ''
                                title_value = title_mapping[excel_title]['title']
                                category_value = title_mapping[excel_title]['category']
                                subcategory_value = title_mapping[excel_title]['subcategory']
                                subcategory_url_value = title_mapping[excel_title]['subcategory_url']
                                
                                # Update Title only if current title is empty or mapping title is more meaningful
                                if not current_title or current_title.strip() == '' or len(title_value.strip()) > len(current_title.strip()):
                                    df.at[idx, 'Title'] = title_value
                                    logger.info(f"DEBUG: Updated Title from '{current_title}' to '{title_value}'")
                                else:
                                    logger.info(f"DEBUG: Keeping original Title: '{current_title}' (mapping title: '{title_value}')")
                                
                                df.at[idx, 'Category'] = category_value
                                df.at[idx, 'Sub-Category'] = subcategory_value
                                df.at[idx, 'Sub-Category Url'] = subcategory_url_value
                                mapped_count += 1
                                found_match = True
                                logger.info(f"DEBUG: Partial match found: '{filename}' matches '{excel_title}' -> Title: '{df.at[idx, 'Title']}', Category: '{category_value}', Sub-Category: '{subcategory_value}', Sub-Category Url: '{subcategory_url_value}'")
                                break
                        
                        if not found_match:
                            logger.info(f"DEBUG: No match found even with partial matching for: '{filename}'")
                
                # Process Schema 1 (Breadcrumb) to update name with Sub-Category and generate URL
                if 'Schema 1' in df.columns:
                    schema1_data = df.at[idx, 'Schema 1']
                    if schema1_data and isinstance(schema1_data, str):
                        try:
                            import json
                            breadcrumb_data = json.loads(schema1_data)
                            
                            # Update position 2 name with Sub-Category and generate URL
                            if breadcrumb_data and isinstance(breadcrumb_data, dict):
                                item_list = breadcrumb_data.get('itemListElement', [])
                                if len(item_list) >= 2:
                                    position_2_item = item_list[1]  # position 2 (0-indexed)
                                    
                                    # Update position 2 name with Sub-Category data
                                    subcategory_data = df.at[idx, 'Sub-Category']
                                    if subcategory_data:
                                        position_2_item['name'] = subcategory_data
                                        logger.info(f"DEBUG: Updated Schema 1 position 2 name with Sub-Category: '{subcategory_data}'")
                                    
                                    # Use Sub-Category URL from mapping instead of generating from SKU
                                    logger.info(f"DEBUG: Checking Sub-Category Url for file: {filename}")
                                    logger.info(f"DEBUG: Available columns: {df.columns.tolist()}")
                                    logger.info(f"DEBUG: Sub-Category Url column exists: {'Sub-Category Url' in df.columns}")
                                    
                                    subcategory_url = df.at[idx, 'Sub-Category Url'] if 'Sub-Category Url' in df.columns else ''
                                    logger.info(f"DEBUG: Sub-Category Url value: '{subcategory_url}'")
                                    
                                    if subcategory_url:
                                        logger.info(f"DEBUG: Using Sub-Category URL: '{subcategory_url}'")
                                        # Update position 2 item with Sub-Category URL
                                        position_2_item['item'] = subcategory_url
                                        
                                        # Update Schema 1 with new URL (formatted)
                                        df.at[idx, 'Schema 1'] = json.dumps(breadcrumb_data, indent=2)
                                        
                                        logger.info(f"DEBUG: Updated Schema 1 with Sub-Category URL: '{subcategory_url}'")
                                    else:
                                        logger.info(f"DEBUG: No Sub-Category URL found, will use fallback logic")
                                        # Fallback: Generate URL from SKU with new processing rules
                                        sku_code = filename  # Use filename as SKU code
                                        logger.info(f"DEBUG: No Sub-Category URL found, generating from SKU: '{sku_code}'")
                                        
                                        if sku_code:
                                            # New SKU processing rules:
                                            # 1. Replace & with space
                                            processed_sku = sku_code.replace('&', ' ')
                                            logger.info(f"DEBUG: After replacing & with space: '{processed_sku}'")
                                            
                                            # 2. Replace - with space  
                                            processed_sku = processed_sku.replace('-', ' ')
                                            logger.info(f"DEBUG: After replacing - with space: '{processed_sku}'")
                                            
                                            # 3. Remove parentheses and content inside, replace with space
                                            processed_sku = re.sub(r'\([^)]*\)', ' ', processed_sku)
                                            logger.info(f"DEBUG: After removing parentheses: '{processed_sku}'")
                                            
                                            # 4. Clean up multiple spaces and trim
                                            processed_sku = re.sub(r'\s+', ' ', processed_sku).strip()
                                            logger.info(f"DEBUG: After cleaning spaces: '{processed_sku}'")
                                            
                                            # 5. Replace remaining spaces with hyphens and convert to lowercase
                                            processed_sku = processed_sku.replace(' ', '-').lower()
                                            
                                            # 6. Append "-market" at the end
                                            processed_sku = f"{processed_sku}-market"
                                            logger.info(f"DEBUG: Final processed SKU: '{processed_sku}'")
                                            
                                            # 7. Generate URL with correct base URL
                                            base_url = "https://www.strategicmarketresearch.com/report"
                                            generated_url = f"{base_url}/{processed_sku}"
                                            
                                            # Update position 2 item with generated URL
                                            position_2_item['item'] = generated_url
                                            
                                            # Update Schema 1 with new URL (formatted)
                                            df.at[idx, 'Schema 1'] = json.dumps(breadcrumb_data, indent=2)
                                            
                                            logger.info(f"DEBUG: Generated URL from SKU '{sku_code}' -> '{generated_url}'")
                                            logger.info(f"DEBUG: Updated Schema 1 with generated URL")
                                        
                        except Exception as e:
                            logger.info(f"DEBUG: Error processing Schema 1 for {filename}: {str(e)}")
            
            logger.info(f"DEBUG: Total files mapped: {mapped_count}/{len(df)}")
            logger.info(f"DEBUG: After mapping - DataFrame columns: {df.columns.tolist()}")
            logger.info(f"DEBUG: After mapping - First few Title values: {df['Title'].head().tolist()}")
            logger.info(f"DEBUG: After mapping - First few Category values: {df['Category'].head().tolist()}")
            logger.info(f"DEBUG: After mapping - First few Sub-Category values: {df['Sub-Category'].head().tolist()}")
            
            # Reorder columns: File, Category, Title, ...other, Schema2, Sub-Category, Description_Part2, Description_Part3
            cols = df.columns.tolist()
            logger.info(f"DEBUG: Before final reordering - columns: {cols}")
            
            # Remove Title, Category, Sub-Category, and Sub-Category Url from current positions
            if 'Title' in cols:
                cols.remove('Title')
            if 'Category' in cols:
                cols.remove('Category')
            if 'Sub-Category' in cols:
                cols.remove('Sub-Category')
            if 'Sub-Category Url' in cols:
                cols.remove('Sub-Category Url')
            
            # Insert Category and Title after File (position 1)
            file_pos = cols.index('File') if 'File' in cols else 0
            cols.insert(file_pos + 1, 'Category')
            cols.insert(file_pos + 2, 'Title')
            logger.info(f"DEBUG: Inserted Category and Title after File at positions {file_pos + 1} and {file_pos + 2}")
            
            # Insert Sub-Category and Sub-Category Url before Description_Part2
            desc_part2_pos = None
            for i, col in enumerate(cols):
                if 'Description_Part2' in col or 'escription_Part2' in col:
                    desc_part2_pos = i
                    break
            
            if desc_part2_pos is not None:
                cols.insert(desc_part2_pos, 'Sub-Category')
                cols.insert(desc_part2_pos + 1, 'Sub-Category Url')
                logger.info(f"DEBUG: Inserted Sub-Category and Sub-Category Url before Description_Part2 at position {desc_part2_pos}")
                logger.info(f"DEBUG: Expected order: File, Category, ..., Schema2, Sub-Category, Sub-Category Url, Description_Part2, Description_Part3")
            elif 'Schema2' in cols:
                # Fallback: insert after Schema2 if Description_Part2 not found
                schema2_pos = cols.index('Schema2')
                cols.insert(schema2_pos + 1, 'Sub-Category')
                cols.insert(schema2_pos + 2, 'Sub-Category Url')
                logger.info(f"DEBUG: Description_Part2 not found, inserted Sub-Category and Sub-Category Url after Schema2 at position {schema2_pos + 1}")
            else:
                # Final fallback: insert at the end
                cols.append('Sub-Category')
                cols.append('Sub-Category Url')
                logger.info(f"DEBUG: Neither Description_Part2 nor Schema2 found, inserted Sub-Category and Sub-Category Url at the end")
            
            logger.info(f"DEBUG: After final reordering - columns: {cols}")
            logger.info(f"DEBUG: Final column order - File, Category, Title, ..., Sub-Category, Sub-Category Url, ...")
            
            # Remove Sub-Category Url from final output (keep for internal processing only)
            if 'Sub-Category Url' in cols:
                cols.remove('Sub-Category Url')
                logger.info(f"DEBUG: Removed Sub-Category Url from final output columns")
            
            df = df[cols]
            logger.info(f"DEBUG: Final DataFrame columns: {df.columns.tolist()}")
            logger.info(f"DEBUG: Final Title values: {df['Title'].head().tolist()}")
            logger.info(f"DEBUG: Final Category values: {df['Category'].head().tolist()}")
            logger.info(f"DEBUG: Final Sub-Category values: {df['Sub-Category'].head().tolist()}")
        
        # Save the updated Excel file with different name
        try:
            import time
            
            # Create new file path with timestamp to avoid cache issues
            timestamp = int(time.time())
            mapped_result_path = result_path.replace('.xlsx', f'_mapped_{timestamp}.xlsx')
            
            # Save mapped file with new name
            df.to_excel(mapped_result_path, index=False, engine='openpyxl')
            
            # Update the result path in JOBS to point to the mapped file
            JOBS[job_id]["result"]["xlsx"] = mapped_result_path
            
            # Also create mapped CSV
            csv_path = result_path.replace('.xlsx', '.csv')
            mapped_csv_path = csv_path.replace('.csv', f'_mapped_{timestamp}.csv')
            if os.path.exists(csv_path):
                df.to_csv(mapped_csv_path, index=False, encoding="utf-8-sig")
                JOBS[job_id]["result"]["csv"] = mapped_csv_path
                
        except Exception as e:
            return Response({"success": False, "message": f"Error saving mapped file: {str(e)}"}, status=500)
        
        return Response({
            "success": True,
            "message": "Excel mapping applied successfully! You can now download the mapped Excel file with Category and Sub-Category data filled."
        })
        
    except Exception as e:
        print(f"ERROR: Apply mapping failed: {str(e)}")
        return Response({"success": False, "message": f"Mapping failed: {str(e)}"}, status=500)

# ------------------- Result download -------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def result_file(request):
    job_id = request.GET.get("jobId")
    if not job_id or job_id not in JOBS:
        raise Http404("job not found")
    if not _job_is_authorized(request, job_id):
        raise Http404("job not found")

    fmt = request.GET.get("format", "xlsx").lower()
    
    
    path = JOBS[job_id].get("result", {}).get(fmt)
    logger.info(f"Path from JOBS: {path}")
    
    # Check if this is a mapped file request and force mapped file path
    if fmt == "xlsx" and JOBS[job_id].get('excel_uploaded', False):
        # Force use mapped file if available
        if path and '_mapped_' not in path:
            # Look for mapped file in the same directory
            import glob
            job_dir = os.path.dirname(path)
            mapped_files = glob.glob(os.path.join(job_dir, "*_mapped_*.xlsx"))
            if mapped_files:
                # Use the most recent mapped file
                mapped_files.sort(key=os.path.getmtime, reverse=True)
                path = mapped_files[0]
                logger.info(f"Forced mapped file path: {path}")
            else:
                logger.warning(f"No mapped file found in directory: {job_dir}")
    
    if not path or not os.path.exists(path):
        logger.warning(f"Path not found or doesn't exist: {path}")
        raise Http404("result not ready")

    # Get the folder name for the download filename
    folder_name = JOBS[job_id].get("folder_name", "Word_Files")
    
    # Add timestamp to filename to avoid cache issues
    import time
    timestamp = int(time.time())
    
    logger.info(f"Download request - job_id={job_id}, folder_name={folder_name}, format={fmt}")
    
    # Mark all files for this job as download complete in database
    UploadedFile.objects.filter(job_id=job_id).update(download_complete=True)
    
    # After marking as downloaded, delete all uploaded files from database
    deleted_count = UploadedFile.objects.filter(job_id=job_id).count()
    UploadedFile.objects.filter(job_id=job_id).delete()
    
    # Also delete Excel mapping data from database
    mapping_deleted = ExcelMapping.objects.filter(job_id=job_id).count()
    ExcelMapping.objects.filter(job_id=job_id).delete()
    
    # Also delete Extract Excel data from database
    extract_deleted = ExtractExcelData.objects.filter(job_id=job_id).count()
    ExtractExcelData.objects.filter(job_id=job_id).delete()
    
    # Delete job record entirely from database
    job_record = JobRecord.objects.filter(job_id=job_id).first()
    if job_record:
        job_record.delete()
        logger.info(f"Deleted job record, deleted {deleted_count} file records, {mapping_deleted} mapping records, and {extract_deleted} extract records from database")
    else:
        logger.warning(f"Job record not found for {job_id}")
    
    if fmt == "csv":
        filename = f"{folder_name}_mapped.csv"
        logger.info(f"DEBUG: Downloading CSV file as: {filename}")  # Debug log
        _remove_job_from_session(request, job_id)
        return FileResponse(open(path, "rb"), as_attachment=True, filename=filename, content_type="text/csv")
    else:
        # Check if Excel mapping was applied
        is_mapped = JOBS[job_id].get('excel_uploaded', False)
        if is_mapped:
            filename = f"{folder_name}_MAPPED.xlsx"
        else:
            filename = f"{folder_name}.xlsx"
        
        logger.info(f"Downloading Excel file as: {filename}")
        
        # Force clean filename without job ID
        if is_mapped:
            clean_filename = f"{folder_name}_MAPPED.xlsx"
        else:
            clean_filename = f"{folder_name}.xlsx"
        
        _remove_job_from_session(request, job_id)
        return FileResponse(open(path, "rb"), as_attachment=True, filename=clean_filename,
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
